import numpy as np
import pandas as pd
from pandas import read_excel
import argparse
import csv
import os
import math
from pandas import DataFrame, read_csv
import xlsxwriter 
from openpyxl import load_workbook

NumToolToCompare = 5         
Tools = ["DSE", "IMP-S", "ARDIFF-R", "ARDIFF-H3","ARDIFF"]
                                      
def stats():
                                                                                                            
                            

    filepath = 'Results.txt'
    file1 = open("TempResults.txt","w") 
    with open(filepath) as fp:
        line = fp.readline()
        while line:
            if "../benchmarks" in line:
                file1.writelines(line)  
            for tool in Tools:
                if "--"+tool+"--" in line:
                    file1.writelines(line)   
            if "-Def-use and uninterpreted functions" in line:
                file1.writelines(line)   
            if "-Symbolic execution" in line:
                file1.writelines(line) 
            if "-Creating Z3 expressions" in line:
                file1.writelines(line)  
            if "-Constraint solving" in line:
                file1.writelines(line) 
            if "Output :"  in line:
                file1.writelines(line)  
            if "-Initialization " in line:
                file1.writelines(line) 
            if "-Program slicing " in line:
                file1.writelines(line)  
            line = fp.readline()
    file1.close() 
                                                                                                            
                        
    Savefile = 'Results.xlsx'
    if (not (os.path.exists(Savefile))):
        workbook = xlsxwriter.Workbook(Savefile)
        worksheet = workbook.add_worksheet("Eq")
        worksheet = workbook.add_worksheet("NEq")
        workbook.close()

    wb = load_workbook(Savefile)
    EqSheet = wb.get_sheet_by_name("Eq")  
    EqSheet.cell(row = 1, column = 1).value = "Benchmark"
    EqSheet.cell(row = 1, column = 2).value = "Method"
    for i in range(0,len(Tools)):
        EqSheet.cell(row = 1, column = 3+(i*2)).value = Tools[i] 
        EqSheet.cell(row = 1, column = 4+(i*2)).value = Tools[i]

    NEqSheet = wb.get_sheet_by_name("NEq")  
    NEqSheet.cell(row = 1, column = 1).value = "Benchmark"
    NEqSheet.cell(row = 1, column = 2).value = "Method"
    for i in range(0,len(Tools)):
        NEqSheet.cell(row = 1, column = 3+(i*2)).value = Tools[i] 
        NEqSheet.cell(row = 1, column = 4+(i*2)).value = Tools[i]
                                                  
    filepath = "TempResults.txt"
    CurrentRowEq = 1
    CurrentRowNEq = 1
    with open(filepath) as fp:
        line = fp.readline()
        while line:
            if "../benchmarks" in line:
                benchmark = line.split("/")[2]
                print("benchmark name: " + benchmark)
                if (benchmark=="ModDiff"):
                    method = line.split("/")[-1].split(" ")[0]
                    case = line.split("/")[-2].split(" ")[0]
                else:
                    method = line.split("/")[-2].split(" ")[0]
                    case = line.split("/")[-1].split(" ")[0] 
                if case == "Eq":
                    CurrentRowEq +=1
                else:
                    CurrentRowNEq +=1

                PreWasUNKOWN = False
                for i in range(0,len(Tools)):
                    if not PreWasUNKOWN:
                        line = fp.readline()              
                    print("tool name:" + line.rstrip())    
                    PreWasUNKOWN = False      

                    line = fp.readline()              
                     
                    if (not line) or line.startswith("####"):                          
                        time = 30000
                        result = "Timeout"

                    elif line.startswith("------"):        
                        PreWasUNKOWN = True
                        time = 30000
                        result = "Timeout"

                    else:             
                        time = 0.0
                        while not line.startswith("Output"):
                            line = line.replace("ms","")
                            time += float(line.split(" : ")[-1].rstrip())
                            line = fp.readline()
                        time = time/1000.0                      
                        stepper = 10.0 ** 3               
                        time = math.trunc(stepper * time) / stepper     
                        result = line.split(" : ")[-1].rstrip()
                        if i==len(Tools)-1:
                           line = fp.readline()                   
                    if case == "Eq":
                        EqSheet.cell(row = CurrentRowEq, column = 1).value = benchmark      
                        EqSheet.cell(row = CurrentRowEq, column = 2).value = method        
                        EqSheet.cell(row = CurrentRowEq, column = 3+(i*2)).value = result             
                        EqSheet.cell(row = CurrentRowEq, column = 4+(i*2)).value = time     
                    else:
                        NEqSheet.cell(row = CurrentRowNEq, column = 1).value = benchmark
                        NEqSheet.cell(row = CurrentRowNEq, column = 2).value = method
                        NEqSheet.cell(row = CurrentRowNEq, column = 3+(i*2)).value = result
                        NEqSheet.cell(row = CurrentRowNEq, column = 4+(i*2)).value = time


                 
                

    wb.save(Savefile) 
    

def main():
    stats()
    
if __name__ == '__main__':
    main()
